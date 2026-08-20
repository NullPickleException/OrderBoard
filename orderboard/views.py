from django.contrib import messages
from django.db import transaction
from django.db.models import Count, F, Q, Sum
from django.db.models.deletion import ProtectedError
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.contrib.auth.decorators import login_required

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, F
from datetime import datetime

from .forms import (
    CustomerContactFormSet,
    CustomerForm,
    OrderEditForm,
    OrderForm,
    OrderPhotoForm,
    PaymentForm,
    ShippingForm,
)

from .models import (
    Customer,
    Order,
    OrderActivity,
    OrderPhoto,
    Payment,
    Shipping,
)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def dashboard(request):
    now = timezone.now()

    orders = (
        Order.objects
        .select_related("customer")
        .prefetch_related("shipping", "payments")
    )

    # =========================================================================
    # ORDER QUEUES
    # =========================================================================

    modeling_orders = (
        orders
        .filter(status=Order.Status.MODELING)
        .order_by("deadline", "-created_at")
    )

    printing_orders = (
        orders
        .filter(status=Order.Status.PRINT)
        .order_by("deadline", "-created_at")
    )

    painting_orders = (
        orders
        .filter(status=Order.Status.PAINT)
        .order_by("deadline", "-created_at")
    )

    completed_orders = (
        orders
        .filter(status=Order.Status.DONE)
        .order_by("-updated_at")
    )

    # =========================================================================
    # DEADLINES
    # =========================================================================

    overdue_orders = (
        orders
        .filter(deadline__lt=now)
        .exclude(
            status__in=[
                Order.Status.DONE,
                Order.Status.CANCELED,
            ]
        )
        .order_by("deadline")[:10]
    )

    upcoming_orders = (
        orders
        .filter(deadline__gte=now)
        .exclude(
            status__in=[
                Order.Status.DONE,
                Order.Status.CANCELED,
            ]
        )
        .order_by("deadline")[:10]
    )

    # =========================================================================
    # STATUS COUNTS
    # =========================================================================

    status_counts = {
        status.value: orders.filter(
            status=status.value
        ).count()
        for status in Order.Status
    }

    # =========================================================================
    # PAYMENTS
    # =========================================================================

    unpaid_orders = [
        order
        for order in orders
        if order.remaining_balance() > 0
    ]

    total_revenue = (
        Payment.objects.aggregate(
            total=Sum("paid_amount")
        )["total"]
        or 0
    )

    recent_payments = (
        Payment.objects
        .select_related(
            "order",
            "order__customer",
        )
        .prefetch_related("order__shipping")
        .order_by("-paid_at")[:10]
    )

    # =========================================================================
    # SHIPPING
    # =========================================================================

    shipping_counts = {
        status.value: Shipping.objects.filter(
            status=status.value
        ).count()
        for status in Shipping.Status
    }

    pending_shipments = (
        Shipping.objects
        .select_related(
            "order",
            "order__customer",
        )
        .exclude(status=Shipping.Status.DELIVERED)
        .exclude(order__status=Order.Status.CANCELED)
        .order_by("order__deadline")[:10]
    )

    not_shipped_orders = (
        Shipping.objects
        .select_related(
            "order",
            "order__customer",
        )
        .filter(status=Shipping.Status.NOT_SHIPPED)
        .exclude(order__status=Order.Status.CANCELED)
        .order_by("order__deadline")
    )

    packaged_orders = (
        Shipping.objects
        .select_related(
            "order",
            "order__customer",
        )
        .filter(status=Shipping.Status.PACKAGED)
        .exclude(order__status=Order.Status.CANCELED)
        .order_by("order__deadline")
    )

    shipped_orders = (
        Shipping.objects
        .select_related(
            "order",
            "order__customer",
        )
        .filter(status=Shipping.Status.SHIPPED)
        .exclude(order__status=Order.Status.CANCELED)
        .order_by("order__deadline")
    )

    delivered_orders = (
        Shipping.objects
        .select_related(
            "order",
            "order__customer",
        )
        .filter(status=Shipping.Status.DELIVERED)
        .exclude(order__status=Order.Status.CANCELED)
        .order_by("-delivered_at")
    )

    # =========================================================================
    # CONTEXT
    # =========================================================================

    context = {
        # Summary
        "total_orders": orders.count(),
        "total_customers": Customer.objects.count(),

        # Status counts
        "status_counts": status_counts,

        # Order queues
        "modeling_orders": modeling_orders,
        "printing_orders": printing_orders,
        "painting_orders": painting_orders,
        "completed_orders": completed_orders,

        # Deadlines
        "overdue_orders": overdue_orders,
        "upcoming_orders": upcoming_orders,

        # Payments
        "unpaid_orders": unpaid_orders[:10],
        "all_unpaid_orders": unpaid_orders,
        "total_revenue": total_revenue,
        "recent_payments": recent_payments,

        # Shipping
        "shipping_counts": shipping_counts,
        "pending_shipments": pending_shipments,
        "not_shipped_orders": not_shipped_orders,
        "packaged_orders": packaged_orders,
        "shipped_orders": shipped_orders,
        "delivered_orders": delivered_orders,

        "now": now,
    }

    return render(
        request,
        "orderboard/dashboard.html",
        context,
    )
    
    
    
# =============================================================================
# ORDER VIEWS
# =============================================================================

# -----------------------------------------------------------------------------
# Order List
# -----------------------------------------------------------------------------
@login_required
def order_list(request):
    orders = (
        Order.objects
        .select_related("customer")
        .prefetch_related("payments", "shipping")
    )

    # =========================================================================
    # Search
    # =========================================================================

    search = request.GET.get("search", "").strip()

    if search:
        orders = orders.filter(
            Q(title__icontains=search)
            | Q(customer__name__icontains=search)
            | Q(customer__phone__icontains=search)
        )

    # =========================================================================
    # Status Filter
    # =========================================================================

    status = request.GET.get("status", "").strip()

    if status:
        orders = orders.filter(
            status=status
        )

    # =========================================================================
    # Shipping Status Filter
    # =========================================================================

    shipping_status = request.GET.get("shipping_status", "").strip()

    if shipping_status:
        orders = orders.filter(
            shipping__status=shipping_status
        )

    # =========================================================================
    # Source Filter
    # =========================================================================

    source = request.GET.get("source", "").strip()

    if source:
        orders = orders.filter(
            source__icontains=source
        )

    # =========================================================================
    # Order Received Date Filter
    # =========================================================================

    received_from = request.GET.get(
        "received_from",
        "",
    ).strip()

    received_to = request.GET.get(
        "received_to",
        "",
    ).strip()

    if received_from:
        orders = orders.filter(
            order_received_at__date__gte=received_from
        )

    if received_to:
        orders = orders.filter(
            order_received_at__date__lte=received_to
        )

    # =========================================================================
    # Deadline Date Filter
    # =========================================================================

    deadline_from = request.GET.get(
        "deadline_from",
        "",
    ).strip()

    deadline_to = request.GET.get(
        "deadline_to",
        "",
    ).strip()

    if deadline_from:
        orders = orders.filter(
            deadline__date__gte=deadline_from
        )

    if deadline_to:
        orders = orders.filter(
            deadline__date__lte=deadline_to
        )

    # =========================================================================
    # Deadline State Filter
    # =========================================================================

    deadline_state = request.GET.get(
        "deadline_state",
        "",
    ).strip()

    now = timezone.now()

    if deadline_state == "overdue":
        orders = (
            orders
            .filter(deadline__lt=now)
            .exclude(
                status__in=[
                    Order.Status.DONE,
                    Order.Status.CANCELED,
                ]
            )
            .exclude(shipping__status=Shipping.Status.DELIVERED)
        )

    elif deadline_state == "upcoming":
        orders = (
            orders
            .filter(deadline__gte=now)
            .exclude(
                status__in=[
                    Order.Status.DONE,
                    Order.Status.CANCELED,
                ]
            )
            .exclude(shipping__status=Shipping.Status.DELIVERED)
        )

    elif deadline_state == "no_deadline":
        orders = orders.filter(
            deadline__isnull=True
        )

    # =========================================================================
    # Payment Status Filter
    # =========================================================================

    payment_status = request.GET.get(
        "payment_status",
        "",
    ).strip()

    if payment_status in [
        "not_paid",
        "partially_paid",
        "fully_paid",
    ]:
        orders_list = list(orders)

        if payment_status == "not_paid":
            orders = [
                order
                for order in orders_list
                if order.total_paid() <= 0
            ]

        elif payment_status == "partially_paid":
            orders = [
                order
                for order in orders_list
                if 0 < order.total_paid() < order.total_price
            ]

        elif payment_status == "fully_paid":
            orders = [
                order
                for order in orders_list
                if order.total_paid() >= order.total_price
            ]

    # =========================================================================
    # Sorting
    # =========================================================================

    sort = request.GET.get(
        "sort",
        "newest",
    ).strip()

    if isinstance(orders, list):

        if sort == "newest":
            orders.sort(
                key=lambda x: x.created_at,
                reverse=True,
            )

        elif sort == "oldest":
            orders.sort(
                key=lambda x: x.created_at
            )

        elif sort == "deadline_soonest":
            orders.sort(
                key=lambda x:
                    x.deadline
                    if x.deadline
                    else timezone.datetime.max.replace(
                        tzinfo=timezone.utc
                    )
            )

        elif sort == "deadline_latest":
            orders.sort(
                key=lambda x:
                    x.deadline
                    if x.deadline
                    else timezone.datetime.min.replace(
                        tzinfo=timezone.utc
                    ),
                reverse=True,
            )

        elif sort == "price_high":
            orders.sort(
                key=lambda x: x.total_price,
                reverse=True,
            )

        elif sort == "price_low":
            orders.sort(
                key=lambda x: x.total_price
            )

    else:

        sort_options = {
            "newest": "-created_at",
            "oldest": "created_at",
            "deadline_soonest": "deadline",
            "deadline_latest": "-deadline",
            "price_high": "-total_price",
            "price_low": "total_price",
        }

        orders = orders.order_by(
            sort_options.get(
                sort,
                "-created_at",
            )
        )

    # =========================================================================
    # Context
    # =========================================================================

    context = {
        "orders": orders,
        "status_choices": Order.Status,

        "search": search,

        "selected_status": status,
        "selected_shipping_status": shipping_status,
        "selected_source": source,

        "received_from": received_from,
        "received_to": received_to,

        "deadline_from": deadline_from,
        "deadline_to": deadline_to,

        "deadline_state": deadline_state,
        "payment_status": payment_status,

        "sort": sort,

        "now": timezone.now(),
    }

    return render(
        request,
        "orderboard/orders/list.html",
        context,
    )
    
    
# -----------------------------------------------------------------------------
# Order Detail
# -----------------------------------------------------------------------------

@login_required
def order_detail(request, id):
    order = get_object_or_404(
        Order.objects.select_related("customer").prefetch_related("shipping"),
        id=id,
    )

    payment_form = PaymentForm(
        order=order
    )

    shipping = getattr(
        order,
        "shipping",
        None,
    )

    shipping_form = ShippingForm(
        instance=shipping
    )

    photo_form = OrderPhotoForm()

    activities = (
        order.activities
        .order_by("-created_at")
    )

    payments = (
        order.payments
        .order_by("-paid_at")
    )

    photos = (
        order.photos
        .order_by("-uploaded_at")
    )

    context = {
        "order": order,

        "payment_form": payment_form,

        "shipping": shipping,
        "shipping_form": shipping_form,

        "photo_form": photo_form,

        "activities": activities,
        "payments": payments,
        "photos": photos,
    }

    return render(
        request,
        "orderboard/orders/detail.html",
        context,
    )


# -----------------------------------------------------------------------------
# Create Order
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def order_create(request):

    if request.method == "POST":

        form = OrderForm(
            request.POST
        )

        if form.is_valid():

            order = form.save()

            Shipping.objects.create(
                order=order,
                status=Shipping.Status.NOT_SHIPPED,
                recipient_name=order.customer.name,
                address=order.customer.address,
                postal_code=order.customer.postal_code,
                phone=order.customer.phone,
            )

            OrderActivity.objects.create(
                order=order,
                activity_type=OrderActivity.ActivityType.CREATED,
                message="Order created.",
            )

            messages.success(
                request,
                "Order created successfully.",
            )

            return redirect(
                "order_detail",
                id=order.id,
            )

    else:

        form = OrderForm()

    return render(
        request,
        "orderboard/orders/create.html",
        {
            "form": form,
        },
    )


# -----------------------------------------------------------------------------
# Edit Order
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def order_edit(request, id):

    order = get_object_or_404(
        Order,
        id=id,
    )

    old_status = order.status

    if request.method == "POST":

        form = OrderEditForm(
            request.POST,
            instance=order,
        )

        if form.is_valid():

            order = form.save()

            OrderActivity.objects.create(
                order=order,
                activity_type=OrderActivity.ActivityType.UPDATED,
                message="Order information updated.",
            )

            if old_status != order.status:

                old_display = dict(
                    Order.Status.choices
                ).get(
                    old_status,
                    old_status,
                )

                new_display = dict(
                    Order.Status.choices
                ).get(
                    order.status,
                    order.status,
                )

                OrderActivity.objects.create(
                    order=order,
                    activity_type=OrderActivity.ActivityType.STATUS_CHANGED,
                    message=(
                        f"Status changed from "
                        f"{old_display} to {new_display}."
                    ),
                )

            messages.success(
                request,
                "Order updated successfully.",
            )

            return redirect(
                "order_detail",
                id=order.id,
            )

    else:

        form = OrderEditForm(
            instance=order
        )

    return render(
        request,
        "orderboard/orders/edit.html",
        {
            "form": form,
            "order": order,
        },
    )


# -----------------------------------------------------------------------------
# Delete Order
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def order_delete(request, id):

    order = get_object_or_404(
        Order,
        id=id,
    )

    if request.method != "POST":
        return redirect(
            "order_detail",
            id=order.id,
        )

    try:

        order.delete()

    except ProtectedError:

        messages.error(
            request,
            "This order cannot be deleted because another record depends on it.",
        )

        return redirect(
            "order_detail",
            id=order.id,
        )

    messages.success(
        request,
        "Order deleted successfully.",
    )

    return redirect(
        "order_list"
    )


# -----------------------------------------------------------------------------
# Customer Search
# -----------------------------------------------------------------------------

@login_required
def customer_search(request):

    query = request.GET.get(
        "q",
        "",
    ).strip()

    if not query:
        return JsonResponse(
            [],
            safe=False,
        )

    customers = (
        Customer.objects
        .filter(
            Q(name__icontains=query)
            | Q(phone__icontains=query)
        )
        .order_by("name")[:10]
    )

    results = [
        {
            "id": customer.id,
            "name": customer.name,
            "phone": customer.phone,
        }
        for customer in customers
    ]

    return JsonResponse(
        results,
        safe=False,
    )


# =============================================================================
# PAYMENT VIEWS
# =============================================================================

# -----------------------------------------------------------------------------
# Payment List
# -----------------------------------------------------------------------------

@login_required
def payment_list(request):

    payments = (
        Payment.objects
        .select_related(
            "order",
            "order__customer",
        )
    )

    # =========================================================================
    # Search
    # =========================================================================

    search = request.GET.get("search", "").strip()

    if search:
        payments = payments.filter(
            Q(order__title__icontains=search)
            | Q(order__customer__name__icontains=search)
            | Q(note__icontains=search)
        )

    # =========================================================================
    # Date Filters
    # =========================================================================

    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if date_from:
        payments = payments.filter(
            paid_at__date__gte=date_from
        )

    if date_to:
        payments = payments.filter(
            paid_at__date__lte=date_to
        )

    # =========================================================================
    # Sorting
    # =========================================================================

    sort = request.GET.get("sort", "newest").strip()

    sort_options = {
        "newest": "-paid_at",
        "oldest": "paid_at",
        "amount_high": "-paid_amount",
        "amount_low": "paid_amount",
    }

    payments = payments.order_by(
        sort_options.get(sort, "-paid_at")
    )

    # =========================================================================
    # Total
    # =========================================================================

    total_payments = payments.aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    # =========================================================================
    # Context
    # =========================================================================

    context = {
        "payments": payments,
        "search": search,
        "date_from": date_from,
        "date_to": date_to,
        "sort": sort,
        "total_payments": total_payments,
    }

    return render(
        request,
        "orderboard/payments/list.html",
        context,
    )


# -----------------------------------------------------------------------------
# Create Payment
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def payment_create(request, id):

    order = get_object_or_404(
        Order,
        id=id,
    )

    if request.method == "POST":

        form = PaymentForm(
            request.POST,
            order=order,
        )

        if form.is_valid():

            payment = form.save(
                commit=False
            )

            payment.order = order
            payment.save()

            OrderActivity.objects.create(
                order=order,
                activity_type=OrderActivity.ActivityType.PAYMENT,
                message=(
                    f"Payment of "
                    f"{payment.paid_amount} added."
                ),
            )

            messages.success(
                request,
                "Payment added successfully.",
            )

        else:

            for errors in form.errors.values():

                for error in errors:

                    messages.error(
                        request,
                        error,
                    )

    return redirect(
        "order_detail",
        id=order.id,
    )


# -----------------------------------------------------------------------------
# Edit Payment
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def payment_edit(request, id):

    payment = get_object_or_404(
        Payment.objects.select_related("order"),
        id=id,
    )

    order = payment.order

    if request.method == "POST":

        old_amount = payment.paid_amount

        form = PaymentForm(
            request.POST,
            instance=payment,
            order=order,
        )

        if form.is_valid():

            payment = form.save()

            OrderActivity.objects.create(
                order=order,
                activity_type=OrderActivity.ActivityType.PAYMENT,
                message=(
                    f"Payment updated from "
                    f"{old_amount} to "
                    f"{payment.paid_amount}."
                ),
            )

            messages.success(
                request,
                "Payment updated successfully.",
            )

            return redirect(
                "order_detail",
                id=order.id,
            )

    else:

        form = PaymentForm(
            instance=payment,
            order=order,
        )

    return render(
        request,
        "orderboard/payments/edit.html",
        {
            "form": form,
            "payment": payment,
            "order": order,
        },
    )


# -----------------------------------------------------------------------------
# Delete Payment
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def payment_delete(request, id):

    payment = get_object_or_404(
        Payment.objects.select_related("order"),
        id=id,
    )

    order = payment.order

    if request.method != "POST":
        return redirect(
            "order_detail",
            id=order.id,
        )

    payment.delete()

    OrderActivity.objects.create(
        order=order,
        activity_type=OrderActivity.ActivityType.PAYMENT,
        message=(
            f"Payment of "
            f"{payment.paid_amount} deleted."
        ),
    )

    messages.success(
        request,
        "Payment deleted successfully.",
    )

    return redirect(
        "order_detail",
        id=order.id,
    )


# =============================================================================
# SHIPPING VIEWS
# =============================================================================

# -----------------------------------------------------------------------------
# Shipping List
# -----------------------------------------------------------------------------

@login_required
def shipping_list(request):

    shipments = (
        Shipping.objects
        .select_related(
            "order",
            "order__customer",
        )
        .exclude(
            order__status=Order.Status.CANCELED
        )
    )

    # =========================================================================
    # Search
    # =========================================================================

    search = request.GET.get(
        "search",
        "",
    ).strip()

    if search:
        shipments = shipments.filter(
            Q(order__title__icontains=search)
            | Q(order__customer__name__icontains=search)
            | Q(recipient_name__icontains=search)
            | Q(tracking_id__icontains=search)
        )

    # =========================================================================
    # Status Filter
    # =========================================================================

    status = request.GET.get(
        "status",
        "",
    ).strip()

    if status:
        shipments = shipments.filter(
            status=status
        )

    # =========================================================================
    # Sorting
    # =========================================================================

    sort = request.GET.get(
        "sort",
        "newest",
    ).strip()

    sort_options = {
        "newest": "-order__created_at",
        "oldest": "order__created_at",
        "deadline_soonest": "order__deadline",
        "recently_shipped": "-shipped_at",
        "recently_delivered": "-delivered_at",
    }

    shipments = shipments.order_by(
        sort_options.get(
            sort,
            "-order__created_at",
        )
    )

    # =========================================================================
    # Context
    # =========================================================================

    context = {
        "shipments": shipments,
        "status_choices": Shipping.Status,
        "search": search,
        "selected_status": status,
        "sort": sort,
        "now": timezone.now(),
    }

    return render(
        request,
        "orderboard/shipping/list.html",
        context,
    )

# -----------------------------------------------------------------------------
# Edit Shipping
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def shipping_edit(request, id):

    # Get the order first
    order = get_object_or_404(
        Order.objects.select_related("customer"),
        id=id,
    )
    
    # Get or create the shipping record for this order
    try:
        shipping = order.shipping
    except Shipping.DoesNotExist:
        # If no shipping record exists, create one
        shipping = Shipping.objects.create(
            order=order,
            status=Shipping.Status.NOT_SHIPPED,
            recipient_name=order.customer.name,
            address=order.customer.address,
            postal_code=order.customer.postal_code,
            phone=order.customer.phone,
        )
    
    old_status = shipping.status

    if request.method == "POST":

        form = ShippingForm(
            request.POST,
            instance=shipping,
        )

        if form.is_valid():

            shipping = form.save()

            OrderActivity.objects.create(
                order=order,
                activity_type=OrderActivity.ActivityType.SHIPPING,
                message="Shipping information updated.",
            )

            if old_status != shipping.status:

                old_display = dict(
                    Shipping.Status.choices
                ).get(
                    old_status,
                    old_status,
                )

                new_display = dict(
                    Shipping.Status.choices
                ).get(
                    shipping.status,
                    shipping.status,
                )

                OrderActivity.objects.create(
                    order=order,
                    activity_type=OrderActivity.ActivityType.SHIPPING,
                    message=(
                        f"Shipping status changed from "
                        f"{old_display} to {new_display}."
                    ),
                )

            messages.success(
                request,
                "Shipping information updated successfully.",
            )

            return redirect(
                "order_detail",
                id=order.id,
            )

    else:

        form = ShippingForm(
            instance=shipping
        )

    return render(
        request,
        "orderboard/shipping/edit.html",
        {
            "form": form,
            "shipping": shipping,
            "order": order,
        },
    )

# =============================================================================
# PHOTO VIEWS
# =============================================================================

# -----------------------------------------------------------------------------
# Photo Upload
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def photo_upload(request, id):

    order = get_object_or_404(
        Order,
        id=id,
    )

    if request.method == "POST":

        form = OrderPhotoForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():

            photo = form.save(
                commit=False
            )

            photo.order = order
            photo.save()

            OrderActivity.objects.create(
                order=order,
                activity_type=OrderActivity.ActivityType.PHOTO,
                message="Order photo uploaded.",
            )

            messages.success(
                request,
                "Photo uploaded successfully.",
            )

        else:

            for errors in form.errors.values():

                for error in errors:

                    messages.error(
                        request,
                        error,
                    )

    return redirect(
        "order_detail",
        id=order.id,
    )


# -----------------------------------------------------------------------------
# Photo Delete
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def photo_delete(request, id):

    photo = get_object_or_404(
        OrderPhoto,
        id=id,
    )

    order = photo.order

    if request.method != "POST":
        return redirect(
            "order_detail",
            id=order.id,
        )

    photo.delete()

    OrderActivity.objects.create(
        order=order,
        activity_type=OrderActivity.ActivityType.PHOTO,
        message="Order photo deleted.",
    )

    messages.success(
        request,
        "Photo deleted.",
    )

    return redirect(
        "order_detail",
        id=order.id,
    )


# =============================================================================
# CUSTOMER VIEWS
# =============================================================================

# -----------------------------------------------------------------------------
# Customer List
# -----------------------------------------------------------------------------

@login_required
def customer_list(request):

    customers = (
        Customer.objects
        .annotate(
            order_count=Count("orders", distinct=True),
            contact_count=Count("contacts", distinct=True),
        )
    )

    # =========================================================================
    # Search
    # =========================================================================

    search = request.GET.get("search", "").strip()

    if search:
        customers = customers.filter(
            Q(name__icontains=search)
            | Q(phone__icontains=search)
            | Q(address__icontains=search)
            | Q(postal_code__icontains=search)
            | Q(contacts__platform__icontains=search)
            | Q(contacts__username__icontains=search)
        ).distinct()

    # =========================================================================
    # Has Orders Filter
    # =========================================================================

    has_orders = request.GET.get("has_orders", "").strip()

    if has_orders == "yes":
        customers = customers.filter(
            order_count__gt=0
        )

    elif has_orders == "no":
        customers = customers.filter(
            order_count=0
        )

    # =========================================================================
    # Has Phone Filter
    # =========================================================================

    has_phone = request.GET.get("has_phone", "").strip()

    if has_phone == "yes":
        customers = customers.exclude(
            Q(phone__isnull=True) | Q(phone="")
        )

    elif has_phone == "no":
        customers = customers.filter(
            Q(phone__isnull=True) | Q(phone="")
        )

    # =========================================================================
    # Has Address Filter
    # =========================================================================

    has_address = request.GET.get("has_address", "").strip()

    if has_address == "yes":
        customers = customers.exclude(
            Q(address__isnull=True) | Q(address="")
        )

    elif has_address == "no":
        customers = customers.filter(
            Q(address__isnull=True) | Q(address="")
        )

    # =========================================================================
    # Has Contacts Filter
    # =========================================================================

    has_contacts = request.GET.get("has_contacts", "").strip()

    if has_contacts == "yes":
        customers = customers.filter(
            contact_count__gt=0
        )

    elif has_contacts == "no":
        customers = customers.filter(
            contact_count=0
        )

    # =========================================================================
    # Created Date Filter
    # =========================================================================

    created_from = request.GET.get("created_from", "").strip()
    created_to = request.GET.get("created_to", "").strip()

    if created_from:
        customers = customers.filter(
            created_at__date__gte=created_from
        )

    if created_to:
        customers = customers.filter(
            created_at__date__lte=created_to
        )

    # =========================================================================
    # Sorting
    # =========================================================================

    sort = request.GET.get("sort", "name").strip()

    sort_options = {
        "name": "name",
        "name_desc": "-name",
        "newest": "-created_at",
        "oldest": "created_at",
        "most_orders": "-order_count",
        "least_orders": "order_count",
        "recently_updated": "-updated_at",
    }

    customers = customers.order_by(
        sort_options.get(sort, "name")
    )

    # =========================================================================
    # Context
    # =========================================================================

    context = {
        "customers": customers,
        "search": search,
        "has_orders": has_orders,
        "has_phone": has_phone,
        "has_address": has_address,
        "has_contacts": has_contacts,
        "created_from": created_from,
        "created_to": created_to,
        "sort": sort,
    }

    return render(
        request,
        "orderboard/customers/list.html",
        context,
    )

# -----------------------------------------------------------------------------
# Customer Detail
# -----------------------------------------------------------------------------

@login_required
def customer_detail(request, id):

    customer = get_object_or_404(
        Customer,
        id=id,
    )

    orders = (
        customer.orders
        .select_related("customer")
        .prefetch_related("shipping")
        .order_by("-created_at")
    )

    contacts = (
        customer.contacts
        .order_by(
            "platform",
            "username",
        )
    )

    return render(
        request,
        "orderboard/customers/detail.html",
        {
            "customer": customer,
            "orders": orders,
            "contacts": contacts,
        },
    )


# -----------------------------------------------------------------------------
# Customer Create
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def customer_create(request):

    if request.method == "POST":

        form = CustomerForm(
            request.POST
        )

        contact_formset = CustomerContactFormSet(
            request.POST
        )

        if form.is_valid() and contact_formset.is_valid():

            customer = form.save()

            contact_formset.instance = customer
            contact_formset.save()

            messages.success(
                request,
                "Customer created successfully.",
            )

            return redirect(
                "customer_detail",
                id=customer.id,
            )

    else:

        form = CustomerForm()
        contact_formset = CustomerContactFormSet()

    return render(
        request,
        "orderboard/customers/create.html",
        {
            "form": form,
            "contact_formset": contact_formset,
        },
    )


# -----------------------------------------------------------------------------
# Customer Edit
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def customer_edit(request, id):

    customer = get_object_or_404(
        Customer,
        id=id,
    )

    if request.method == "POST":

        form = CustomerForm(
            request.POST,
            instance=customer,
        )

        contact_formset = CustomerContactFormSet(
            request.POST,
            instance=customer,
        )

        if form.is_valid() and contact_formset.is_valid():

            form.save()
            contact_formset.save()

            messages.success(
                request,
                "Customer updated successfully.",
            )

            return redirect(
                "customer_detail",
                id=customer.id,
            )

    else:

        form = CustomerForm(
            instance=customer,
        )

        contact_formset = CustomerContactFormSet(
            instance=customer,
        )

    return render(
        request,
        "orderboard/customers/edit.html",
        {
            "form": form,
            "contact_formset": contact_formset,
            "customer": customer,
        },
    )


# -----------------------------------------------------------------------------
# Customer Delete
# -----------------------------------------------------------------------------

@login_required
@transaction.atomic
def customer_delete(request, id):

    customer = get_object_or_404(
        Customer,
        id=id,
    )

    if request.method != "POST":
        return redirect(
            "customer_list"
        )

    try:

        customer.delete()

    except ProtectedError:

        messages.error(
            request,
            "This customer cannot be deleted because they have existing orders.",
        )

        return redirect(
            "customer_list"
        )

    messages.success(
        request,
        "Customer deleted successfully.",
    )

    return redirect(
        "customer_list"
    )


# =============================================================================
# PRINTING
# =============================================================================

@login_required
def printing_list(request):
    shipments = (
        Shipping.objects
        .select_related(
            "order",
            "order__customer",
        )
        .exclude(
            order__status=Order.Status.CANCELED
        )
        .order_by(
            "order__deadline",
            "-order__created_at",
        )
    )

    context = {
        "shipments": shipments,

        # Sender information.
        # These are initial values only.
        # The user can modify them directly in the printing page.
        "sender_name": "فلان فلانی",
        "sender_phone": "123456789",
        "sender_address": "",
    }

    return render(
        request,
        "orderboard/printing/list.html",
        context,
    )


# EXPORT VIEWS
# =============================================================================

# -----------------------------------------------------------------------------
# Export Center
# -----------------------------------------------------------------------------

@login_required
def export_center(request):
    """Display the export configuration page."""
    
    context = {
        "order_status_choices": Order.Status.choices,
        "shipping_status_choices": Shipping.Status.choices,
        "export_fields": {
            "orders": {
                "fields": {
                    "id": "Order ID",
                    "customer_name": "Customer Name",
                    "customer_phone": "Customer Phone",
                    "title": "Title",
                    "description": "Description",
                    "source": "Source",
                    "status": "Status",
                    "order_received_at": "Received At",
                    "deadline": "Deadline",
                    "total_price": "Total Price",
                    "project_cost": "Project Cost",
                    "cost_details": "Cost Details",
                    "notes": "Notes",
                    "total_paid": "Total Paid",
                    "remaining_balance": "Remaining Balance",
                    "payment_status": "Payment Status",
                    "created_at": "Created At",
                    "updated_at": "Updated At",
                }
            },
            "customers": {
                "fields": {
                    "id": "Customer ID",
                    "name": "Name",
                    "phone": "Phone",
                    "address": "Address",
                    "postal_code": "Postal Code",
                    "order_count": "Order Count",
                    "contact_count": "Contact Count",
                    "total_spent": "Total Spent",
                    "created_at": "Created At",
                    "updated_at": "Updated At",
                }
            },
            "payments": {
                "fields": {
                    "id": "Payment ID",
                    "order_id": "Order ID",
                    "order_title": "Order Title",
                    "customer_name": "Customer Name",
                    "paid_amount": "Paid Amount",
                    "note": "Note",
                    "paid_at": "Paid At",
                }
            },
            "shipping": {
                "fields": {
                    "id": "Shipping ID",
                    "order_id": "Order ID",
                    "order_title": "Order Title",
                    "customer_name": "Customer Name",
                    "status": "Status",
                    "tracking_id": "Tracking ID",
                    "recipient_name": "Recipient Name",
                    "address": "Address",
                    "postal_code": "Postal Code",
                    "phone": "Phone",
                    "shipped_at": "Shipped At",
                    "delivered_at": "Delivered At",
                }
            },
        },
    }
    
    return render(request, "orderboard/export/index.html", context)


# -----------------------------------------------------------------------------
# Export Excel
# -----------------------------------------------------------------------------

@login_required
def export_excel(request):
    """Generate and download Excel file with filtered data."""
    
    if request.method != "POST":
        return redirect("export_center")
    
    # Get selected datasets
    datasets = request.POST.getlist("datasets", [])
    
    if not datasets:
        messages.error(request, "Please select at least one dataset to export.")
        return redirect("export_center")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Export each selected dataset
    for dataset in datasets:
        if dataset == "orders":
            _export_orders_sheet(wb, request)
        elif dataset == "customers":
            _export_customers_sheet(wb, request)
        elif dataset == "payments":
            _export_payments_sheet(wb, request)
        elif dataset == "shipping":
            _export_shipping_sheet(wb, request)
    
    # If no sheets were created (all exports failed), create a default
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Empty")
    
    # Prepare response
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"orderboard_export_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


# =============================================================================
# EXPORT HELPER FUNCTIONS
# =============================================================================

def _apply_sheet_styling(ws, headers, data_rows):
    """Apply consistent styling to a worksheet."""
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data styling
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    # Border styling
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        
        for row_data in data_rows:
            if col_idx - 1 < len(row_data):
                cell_value = str(row_data[col_idx - 1])
                if len(cell_value) > max_length:
                    max_length = min(len(cell_value), 50)
        
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data_rows) + 1}"


def _export_orders_sheet(wb, request):
    """Export orders data to Excel sheet."""
    
    selected_fields = request.POST.getlist("order_fields", [])
    if not selected_fields:
        selected_fields = ["id", "customer_name", "title", "status", "total_price", "deadline"]
    
    orders = Order.objects.select_related("customer").prefetch_related("payments")
    
    search = request.POST.get("order_search", "").strip()
    if search:
        orders = orders.filter(
            Q(title__icontains=search)
            | Q(customer__name__icontains=search)
            | Q(customer__phone__icontains=search)
        )
    
    status = request.POST.get("order_status", "").strip()
    if status:
        orders = orders.filter(status=status)
    
    shipping_status = request.POST.get("order_shipping_status", "").strip()
    if shipping_status:
        orders = orders.filter(shipping__status=shipping_status)
    
    source = request.POST.get("order_source", "").strip()
    if source:
        orders = orders.filter(source__icontains=source)
    
    received_from = request.POST.get("order_received_from", "").strip()
    if received_from:
        orders = orders.filter(order_received_at__date__gte=received_from)
    
    received_to = request.POST.get("order_received_to", "").strip()
    if received_to:
        orders = orders.filter(order_received_at__date__lte=received_to)
    
    deadline_from = request.POST.get("order_deadline_from", "").strip()
    if deadline_from:
        orders = orders.filter(deadline__date__gte=deadline_from)
    
    deadline_to = request.POST.get("order_deadline_to", "").strip()
    if deadline_to:
        orders = orders.filter(deadline__date__lte=deadline_to)
    
    deadline_state = request.POST.get("order_deadline_state", "").strip()
    now = timezone.now()
    
    if deadline_state == "overdue":
        orders = orders.filter(deadline__lt=now).exclude(
            status__in=[Order.Status.DONE, Order.Status.CANCELED]
        ).exclude(shipping__status=Shipping.Status.DELIVERED)
    elif deadline_state == "upcoming":
        orders = orders.filter(deadline__gte=now).exclude(
            status__in=[Order.Status.DONE, Order.Status.CANCELED]
        ).exclude(shipping__status=Shipping.Status.DELIVERED)
    elif deadline_state == "no_deadline":
        orders = orders.filter(deadline__isnull=True)
    
    payment_status = request.POST.get("order_payment_status", "").strip()
    if payment_status in ["not_paid", "partially_paid", "fully_paid"]:
        orders_list = list(orders)
        if payment_status == "not_paid":
            orders = [o for o in orders_list if o.total_paid() <= 0]
        elif payment_status == "partially_paid":
            orders = [o for o in orders_list if 0 < o.total_paid() < o.total_price]
        elif payment_status == "fully_paid":
            orders = [o for o in orders_list if o.total_paid() >= o.total_price]
    
    field_mapping = {
        "id": lambda o: o.id,
        "customer_name": lambda o: o.customer.name,
        "customer_phone": lambda o: o.customer.phone,
        "title": lambda o: o.title,
        "description": lambda o: o.description,
        "source": lambda o: o.source,
        "status": lambda o: o.get_status_display(),
        "order_received_at": lambda o: o.order_received_at.strftime("%Y-%m-%d %H:%M") if o.order_received_at else "",
        "deadline": lambda o: o.deadline.strftime("%Y-%m-%d %H:%M") if o.deadline else "",
        "total_price": lambda o: float(o.total_price),
        "project_cost": lambda o: float(o.project_cost) if o.project_cost else "",
        "cost_details": lambda o: o.cost_details,
        "notes": lambda o: o.notes,
        "total_paid": lambda o: float(o.total_paid()),
        "remaining_balance": lambda o: float(o.remaining_balance()),
        "payment_status": lambda o: o.payment_status(),
        "created_at": lambda o: o.created_at.strftime("%Y-%m-%d %H:%M"),
        "updated_at": lambda o: o.updated_at.strftime("%Y-%m-%d %H:%M"),
    }
    
    field_labels = {
        "id": "Order ID", "customer_name": "Customer Name", "customer_phone": "Customer Phone",
        "title": "Title", "description": "Description", "source": "Source", "status": "Status",
        "order_received_at": "Received At", "deadline": "Deadline", "total_price": "Total Price",
        "project_cost": "Project Cost", "cost_details": "Cost Details", "notes": "Notes",
        "total_paid": "Total Paid", "remaining_balance": "Remaining Balance",
        "payment_status": "Payment Status", "created_at": "Created At", "updated_at": "Updated At",
    }
    
    ws = wb.create_sheet("Orders")
    headers = [field_labels[f] for f in selected_fields if f in field_mapping]
    
    data_rows = []
    for order in orders:
        row = []
        for field in selected_fields:
            if field in field_mapping:
                try:
                    row.append(field_mapping[field](order))
                except:
                    row.append("")
        data_rows.append(row)
    
    _apply_sheet_styling(ws, headers, data_rows)


def _export_customers_sheet(wb, request):
    """Export customers data to Excel sheet."""
    
    selected_fields = request.POST.getlist("customer_fields", [])
    if not selected_fields:
        selected_fields = ["id", "name", "phone", "address", "order_count", "total_spent"]
    
    customers = Customer.objects.annotate(
        order_count=Count("orders", distinct=True),
        contact_count=Count("contacts", distinct=True),
    )
    
    search = request.POST.get("customer_search", "").strip()
    if search:
        customers = customers.filter(
            Q(name__icontains=search) | Q(phone__icontains=search) | Q(address__icontains=search)
            | Q(postal_code__icontains=search) | Q(contacts__platform__icontains=search)
            | Q(contacts__username__icontains=search)
        ).distinct()
    
    has_orders = request.POST.get("customer_has_orders", "").strip()
    if has_orders == "yes":
        customers = customers.filter(order_count__gt=0)
    elif has_orders == "no":
        customers = customers.filter(order_count=0)
    
    has_phone = request.POST.get("customer_has_phone", "").strip()
    if has_phone == "yes":
        customers = customers.exclude(Q(phone__isnull=True) | Q(phone=""))
    elif has_phone == "no":
        customers = customers.filter(Q(phone__isnull=True) | Q(phone=""))
    
    has_address = request.POST.get("customer_has_address", "").strip()
    if has_address == "yes":
        customers = customers.exclude(Q(address__isnull=True) | Q(address=""))
    elif has_address == "no":
        customers = customers.filter(Q(address__isnull=True) | Q(address=""))
    
    has_contacts = request.POST.get("customer_has_contacts", "").strip()
    if has_contacts == "yes":
        customers = customers.filter(contact_count__gt=0)
    elif has_contacts == "no":
        customers = customers.filter(contact_count=0)
    
    created_from = request.POST.get("customer_created_from", "").strip()
    if created_from:
        customers = customers.filter(created_at__date__gte=created_from)
    
    created_to = request.POST.get("customer_created_to", "").strip()
    if created_to:
        customers = customers.filter(created_at__date__lte=created_to)
    
    field_mapping = {
        "id": lambda c: c.id,
        "name": lambda c: c.name,
        "phone": lambda c: c.phone,
        "address": lambda c: c.address,
        "postal_code": lambda c: c.postal_code,
        "order_count": lambda c: c.order_count,
        "contact_count": lambda c: c.contact_count,
        "total_spent": lambda c: float(c.orders.aggregate(total=Sum("total_price"))["total"] or 0),
        "created_at": lambda c: c.created_at.strftime("%Y-%m-%d %H:%M"),
        "updated_at": lambda c: c.updated_at.strftime("%Y-%m-%d %H:%M"),
    }
    
    field_labels = {
        "id": "Customer ID", "name": "Name", "phone": "Phone", "address": "Address",
        "postal_code": "Postal Code", "order_count": "Order Count", "contact_count": "Contact Count",
        "total_spent": "Total Spent", "created_at": "Created At", "updated_at": "Updated At",
    }
    
    ws = wb.create_sheet("Customers")
    headers = [field_labels[f] for f in selected_fields if f in field_mapping]
    
    data_rows = []
    for customer in customers:
        row = []
        for field in selected_fields:
            if field in field_mapping:
                try:
                    row.append(field_mapping[field](customer))
                except:
                    row.append("")
        data_rows.append(row)
    
    _apply_sheet_styling(ws, headers, data_rows)


def _export_payments_sheet(wb, request):
    """Export payments data to Excel sheet."""
    
    selected_fields = request.POST.getlist("payment_fields", [])
    if not selected_fields:
        selected_fields = ["id", "order_title", "customer_name", "paid_amount", "paid_at"]
    
    payments = Payment.objects.select_related("order", "order__customer")
    
    search = request.POST.get("payment_search", "").strip()
    if search:
        payments = payments.filter(
            Q(order__title__icontains=search) | Q(order__customer__name__icontains=search) | Q(note__icontains=search)
        )
    
    date_from = request.POST.get("payment_date_from", "").strip()
    if date_from:
        payments = payments.filter(paid_at__date__gte=date_from)
    
    date_to = request.POST.get("payment_date_to", "").strip()
    if date_to:
        payments = payments.filter(paid_at__date__lte=date_to)
    
    field_mapping = {
        "id": lambda p: p.id,
        "order_id": lambda p: p.order.id,
        "order_title": lambda p: p.order.title,
        "customer_name": lambda p: p.order.customer.name,
        "paid_amount": lambda p: float(p.paid_amount),
        "note": lambda p: p.note,
        "paid_at": lambda p: p.paid_at.strftime("%Y-%m-%d %H:%M"),
    }
    
    field_labels = {
        "id": "Payment ID", "order_id": "Order ID", "order_title": "Order Title",
        "customer_name": "Customer Name", "paid_amount": "Paid Amount", "note": "Note", "paid_at": "Paid At",
    }
    
    ws = wb.create_sheet("Payments")
    headers = [field_labels[f] for f in selected_fields if f in field_mapping]
    
    data_rows = []
    for payment in payments:
        row = []
        for field in selected_fields:
            if field in field_mapping:
                try:
                    row.append(field_mapping[field](payment))
                except:
                    row.append("")
        data_rows.append(row)
    
    _apply_sheet_styling(ws, headers, data_rows)


def _export_shipping_sheet(wb, request):
    """Export shipping data to Excel sheet."""
    
    selected_fields = request.POST.getlist("shipping_fields", [])
    if not selected_fields:
        selected_fields = ["id", "order_title", "customer_name", "status", "recipient_name", "tracking_id"]
    
    shipments = Shipping.objects.select_related("order", "order__customer")
    
    search = request.POST.get("shipping_search", "").strip()
    if search:
        shipments = shipments.filter(
            Q(order__title__icontains=search) | Q(order__customer__name__icontains=search)
            | Q(recipient_name__icontains=search) | Q(tracking_id__icontains=search)
        )
    
    status = request.POST.get("shipping_status", "").strip()
    if status:
        shipments = shipments.filter(status=status)
    
    field_mapping = {
        "id": lambda s: s.id,
        "order_id": lambda s: s.order.id,
        "order_title": lambda s: s.order.title,
        "customer_name": lambda s: s.order.customer.name,
        "status": lambda s: s.get_status_display(),
        "tracking_id": lambda s: s.tracking_id,
        "recipient_name": lambda s: s.recipient_name,
        "address": lambda s: s.address,
        "postal_code": lambda s: s.postal_code,
        "phone": lambda s: s.phone,
        "shipped_at": lambda s: s.shipped_at.strftime("%Y-%m-%d %H:%M") if s.shipped_at else "",
        "delivered_at": lambda s: s.delivered_at.strftime("%Y-%m-%d %H:%M") if s.delivered_at else "",
    }
    
    field_labels = {
        "id": "Shipping ID", "order_id": "Order ID", "order_title": "Order Title",
        "customer_name": "Customer Name", "status": "Status", "tracking_id": "Tracking ID",
        "recipient_name": "Recipient Name", "address": "Address", "postal_code": "Postal Code",
        "phone": "Phone", "shipped_at": "Shipped At", "delivered_at": "Delivered At",
    }
    
    ws = wb.create_sheet("Shipping")
    headers = [field_labels[f] for f in selected_fields if f in field_mapping]
    
    data_rows = []
    for shipment in shipments:
        row = []
        for field in selected_fields:
            if field in field_mapping:
                try:
                    row.append(field_mapping[field](shipment))
                except:
                    row.append("")
        data_rows.append(row)
    
    _apply_sheet_styling(ws, headers, data_rows)